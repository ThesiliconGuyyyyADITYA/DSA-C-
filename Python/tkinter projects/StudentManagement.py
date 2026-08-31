from tkinter import *

from openpyxl import Workbook, load_workbook
import os

screen = Tk()

screen.geometry("2000x2000")

screen.title("Student Management System")

screen.config(background = "Black")

for i in range(6):
    screen.grid_columnconfigure(i, weight=1)

label = Label(screen , 
            text = "Manage student data" , 
            font = ('Times New Roman' , 28 , 'bold') , 
            fg = "white" ,
            relief = "raised" , 
            bd = 5 , 
            padx = 10 , 
            pady = 10 , 
            )

label.grid(row=0, column=1, columnspan=4)

Label(screen, text="Name").grid(row=2, column=1)

name_entry = Entry(screen)
name_entry.grid(row=2, column=2)

Label(screen, text="Class").grid(row=3, column=1)

class_entry = Entry(screen)
class_entry.grid(row=3, column=2)

Label(screen, text="Roll Number").grid(row=4, column=1)

roll_entry = Entry(screen)
roll_entry.grid(row=4, column=2)

Label(screen, text="School Name").grid(row=5, column=1)

school_entry = Entry(screen)
school_entry.grid(row=5, column=2)

def add_student():

    name = name_entry.get()
    student_class = class_entry.get()
    roll = roll_entry.get()
    school = school_entry.get()
    filename = "students.xlsx"
    print("Saving to:", filename)
    
    if not os.path.exists(filename):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
        "Name",
        "Class",
        "Roll Number",
        "School Name"
        ])

    else:

        workbook = load_workbook(filename)
        sheet = workbook.active

    print("Student Name is" , name)
    print("Class is" , student_class)
    print("Roll Number is" , roll)
    print("School is" , school)
    name_entry.delete(0, END)
    class_entry.delete(0, END)
    roll_entry.delete(0, END)
    school_entry.delete(0, END)

    sheet.append([
    name,
    student_class,
    roll,
    school
    ])

    print("Row added")

    workbook.save(filename)

    print("Workbook saved")

Button(screen , 
       text="Add Student" , 
       command = add_student).grid(
        row=1, column=0)
       
Label(screen, text="Roll Number to Delete").grid(row=7, column=1)

delete_roll_entry = Entry(screen)
delete_roll_entry.grid(row=7, column=2)
       
screen.mainloop()